import os
import statistics
import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import Tk, filedialog, simpledialog

def select_excel_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Excel File", filetypes=[("Excel files", "*.xlsx")]
    )
    return file_path

def get_thresholds():
    try:
        min_spc = float(simpledialog.askstring("Minimum SpC", "Enter minimum SpC value:"))
        min_tic = float(simpledialog.askstring("Minimum TIC", "Enter minimum TIC value:"))
        return min_spc, min_tic
    except (TypeError, ValueError):
        print("Invalid threshold input.")
        return None, None

def get_output_filename():
    filename = simpledialog.askstring(
        "Output File Name",
        "Enter filename for significantly changed proteins (without extension):"
    )
    return f"{filename}.txt" if filename else None

def load_workbook_and_identify_columns(file_path):
    wb = openpyxl.load_workbook(file_path)
    if "Protein list" not in wb.sheetnames:
        print("Sheet 'Protein list' not found.")
        return None, None, None, None

    sheet = wb["Protein list"]
    headers = [cell.value for cell in sheet[1]]

    try:
        protein_col = headers.index("ProteinAC")
    except ValueError:
        try:
            protein_col = headers.index("Protein AC")
        except ValueError:
            print("Column 'ProteinAC' or 'Protein AC' not found.")
            return None, None, None, None

    spc_cols = [i for i, h in enumerate(headers) if isinstance(h, str) and h.startswith("SpC")]
    tic_cols = [i for i, h in enumerate(headers) if isinstance(h, str) and h.startswith("TIC")]

    if not spc_cols and not tic_cols:
        print("No 'SpC' or 'TIC' columns found.")
        return None, None, None, None
    return wb, sheet, protein_col, (spc_cols, tic_cols)

def write_summary_sheet(wb, data_matrix):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.append(["Protein AC", "Average SpC", "Average TIC"])
    for row in data_matrix:
        summary_sheet.append(row)

def main():
    file_path = select_excel_file()
    if not file_path:
        print("No file selected.")
        return

    min_spc, min_tic = get_thresholds()
    if min_spc is None or min_tic is None:
        return

    output_txt_file = get_output_filename()
    if not output_txt_file:
        print("No output filename provided.")
        return

    wb, sheet, protein_col, (spc_cols, tic_cols) = load_workbook_and_identify_columns(file_path)
    if not wb:
        return

    data_matrix = []
    most_changed_proteins = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        protein_ac = row[protein_col]
        if not protein_ac:
            continue

        spc_vals = [row[i] for i in spc_cols if isinstance(row[i], (int, float))]
        tic_vals = [row[i] for i in tic_cols if isinstance(row[i], (int, float))]

        avg_spc = round(statistics.mean(spc_vals), 2) if spc_vals else 0.0
        avg_tic = round(statistics.mean(tic_vals), 2) if tic_vals else 0.0

        data_matrix.append([protein_ac, avg_spc, avg_tic])

        if avg_spc >= min_spc and avg_tic >= min_tic:
            most_changed_proteins.append(protein_ac)

    write_summary_sheet(wb, data_matrix)
    wb.save(file_path)

    with open(output_txt_file, "w") as f:
        for protein in most_changed_proteins:
            f.write(protein + "\n")

    print(f"\nSummary written to 'Summary' sheet in {os.path.basename(file_path)}")
    print(f"Significantly changed proteins saved to {output_txt_file}")

if __name__ == "__main__":
    main()
